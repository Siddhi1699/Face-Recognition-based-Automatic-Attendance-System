import openpyxl
from datetime import date

def setDefault(date,subject):
    path = 'G:\\College\\Study\\6thSem\\STL\\Attendance.xlsx'
    wb = openpyxl.load_workbook(path)
    if subject == 1:
        sheet = wb['AI']
    elif subject == 2:
        sheet = wb['CG']
    elif subject == 3:
        sheet = wb['DBMS']
    elif subject == 4:
        sheet = wb['ADS']

    for i in range(2,sheet.max_row+1):
        c = sheet.cell(row=i, column=sheet.max_column).value = 'A'
    wb.save('G:\\College\\Study\\6thSem\\STL\\Attendance.xlsx')


def update(date,subject,id):
    path = 'G:\\College\\Study\\6thSem\\STL\\Attendance.xlsx'
    wb = openpyxl.load_workbook(path)
    if subject==1:
        sheet = wb['AI']
    elif subject==2:
        sheet = wb['CG']
    elif subject == 3:
        sheet = wb['DBMS']
    elif subject == 4:
        sheet = wb['ADS']

    if sheet.cell(row=1,column=sheet.max_column).value==date:
        c=sheet.cell(row=id+1, column=sheet.max_column).value = 'P'
    else:
        d=sheet.cell(row=1, column=sheet.max_column+1).value = date
        c=sheet.cell(row=id+1, column=sheet.max_column).value = 'P'
    wb.save('G:\\College\\Study\\6thSem\\STL\\Attendance.xlsx')


'''today = date.today()
setDefault(today.strftime("%d/%m/%Y"),1)
update(today.strftime("%d/%m/%Y"),1,2)'''