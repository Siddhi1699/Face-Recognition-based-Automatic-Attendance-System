import openpyxl
import tkinter as tk
#students=[["Manjiri","Vairagade","mv@gmail.com"],["Shreya","Rathi","sr@gmail.com"],["Siddhi","Belgamwar","sb@gmail.com"]]

path = "G:\\College\\Study\\6thSem\\STL\\Attendance.xlsx"

def read(name):
    r=tk.Tk()
    r.title(name)
    r.geometry("500x200")

    def perc(sub_name,n):
        wb_obj = openpyxl.load_workbook(path)
        #sheet_obj = wb_obj.active
        sheet_obj = wb_obj.get_sheet_by_name(sub_name)
        r = sheet_obj.max_row
        c = sheet_obj.max_column
        for i in range(1, r + 1):
            if sheet_obj.cell(row=i, column=1).value == name:
                if n==1:
                    return str(sheet_obj.cell(row=i, column=4).value)
                elif n==2:
                    return str(sheet_obj.cell(row=i, column=5).value)
                elif n==3:
                    return str(sheet_obj.cell(row=i, column=6).value)
                elif n==4:
                    return str(sheet_obj.cell(row=i, column=7).value)
                else:
                    return str(sheet_obj.cell(row=i, column=8).value)


    tk.Label(r,text="Subject",font=("Helvetica", 16)).grid(row=0,column=0)
    tk.Label(r, text="Attendance",font=("Helvetica", 16)).grid(row=0, column=1)
    tk.Label(r, text="ADS",font=("Times", 14)).grid(row=1, column=0)
    tk.Label(r, text="DBMS",font=("Times", 14)).grid(row=2, column=0)
    tk.Label(r, text="AI",font=("Times", 14)).grid(row=3, column=0)
    tk.Label(r, text="CG",font=("Times", 14)).grid(row=4, column=0)

    tk.Label(r, text="%s"%perc("Record",1),font=("Times", 14)).grid(row=1, column=1)
    tk.Label(r, text="%s"%perc("Record",4),font=("Times", 14)).grid(row=2, column=1)
    tk.Label(r, text="%s"%perc("Record",2),font=("Times", 14)).grid(row=3, column=1)
    tk.Label(r, text="%s"%perc("Record",3),font=("Times", 14)).grid(row=4, column=1)

    tk.Label(r, text="\tTotal attendance: %s "%perc("Record",0),font=("Helvetica", 16)).grid(row=7, column=0)


    r.mainloop()


              #  print(sheet_obj.cell(row=i, column=j).value)





    '''for i in students:
        r = sheet.max_row + 1
        for j in range(3):
            c2 = sheet.cell(row=r, column=j + 1)
            c2.value = i[j]
        # sheet.cell(sheet.max_row + 1, 1).value(students[j])
        # sheet.cell(sheet.max_row + 1, 1).value(students[j])
    wb.save("C:\\Users\\Siddhi\\Desktop\\frbas.xlsx")
    # wb_obj = openpyxl.load_workbook(path)
    # sheet_obj = wb_obj.active
    # read again'''
    '''for i in range(1,sheet.max_column+1):
        print(sheet.cell(row=1, column=i).value)'''